import pandas as pd
import numpy as np
import openpyxl as opx
from openpyxl.styles import Alignment, Font, PatternFill, Color, Border, Side
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import datetime
import locale
from tkinter import filedialog
from tkinter import *

root = Tk(className= '   DAGPLANNINGMAKER')
w=500
h=100
ws=root.winfo_screenwidth()
hs=root.winfo_screenheight()
x=(ws/2)-(w/2)
y=(hs/2)-(h/2)
root.geometry('%dx%d+%d+%d'%(w,h,x,y))
root.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("xlsx files","*. xlsx"),("all files","*.*")))
exportlocatie = root.filename

# INLEZEN VAN SAP DATA
# KOLOMMEN VERWIJDEREN, TOEVOEGEN
# ORDENEN VOLGENS ROUTENUMMER EN KOPPELEN VOLGENS NAAM

df = pd.read_excel(exportlocatie)
df = df.drop(['Orderdatum', 'Infotekst', 'Activiteit', 'Afvalfractie', 'Voertuig 2', 'Redencode', 'Omschr. redencode', 'Werknemer 4'], axis=1)
df.dropna(thresh=3, inplace = True)
df.rename(columns = {'Werknemer 1':'Chauffeur', 'Voertuig 1': 'Voertuig', 'Werknemer 2': 'Milieuwerker 1', 'Werknemer 3': 'Milieuwerker 2'}, inplace = True)
df['new'] = df.groupby('Chauffeur')['Route'].transform('max')
df = df.sort_values(by=['new', 'Chauffeur','Route'], ascending=[True, True, False]).drop('new', axis=1)
df['Opmerkingen'] = np.where(df['Route'].str.contains("HHPMD"), "START PMD: ", "")
df['OK/NOK'] = ""
df['Gedaan'] = ""
df['Binnen'] = ""

# GEWENSTE VOLGORDE VAN DE FRACTIES

huisvuil = df[df['Route'].str.contains("HHRESZ|HHPMD|HHRESK")]
diftar = df[df['Route'].str.contains("HHGFZC|HHGFZZ|HHRESC")]
papier = df[df['Route'].str.contains("HHPAP")]
glas = df[df['Route'].str.contains("HHGLS")]
grof = df[df['Route'].str.contains("HHGV|HHGHSV")]
bedrijf = df[df['Route'].str.contains("BA")]
trucks = df[df['Route'].str.contains("AA")]
ondergronds = df[df['Route'].str.contains("OC")]

# NIEUWE RIJEN TOEVOEGEN VOOR HEADERS

new_row = pd.DataFrame({'Voertuig':'RESTAFVAL/PMD', 'Route':'', 'Chauffeur':'', 'Milieuwerker 1':'', 'Milieuwerker 2':'', 'Opmerkingen':'', 'OK/NOK':'','Gedaan':'', 'Binnen':''}, index=[0])
new_row2 = pd.DataFrame({'Voertuig':'DIFTAR', 'Route':'', 'Chauffeur':'', 'Milieuwerker 1':'', 'Milieuwerker 2':'', 'Opmerkingen':'', 'OK/NOK':'','Gedaan':'', 'Binnen':''}, index=[0])
new_row3 = pd.DataFrame({'Voertuig':'PAPIER', 'Route':'', 'Chauffeur':'', 'Milieuwerker 1':'', 'Milieuwerker 2':'', 'Opmerkingen':'', 'OK/NOK':'','Gedaan':'', 'Binnen':''}, index=[0])
new_row4 = pd.DataFrame({'Voertuig':'GLAS', 'Route':'', 'Chauffeur':'', 'Milieuwerker 1':'', 'Milieuwerker 2':'', 'Opmerkingen':'', 'OK/NOK':'','Gedaan':'', 'Binnen':''}, index=[0])
new_row5 = pd.DataFrame({'Voertuig':'GROF VUIL', 'Route':'', 'Chauffeur':'', 'Milieuwerker 1':'', 'Milieuwerker 2':'', 'Opmerkingen':'', 'OK/NOK':'','Gedaan':'', 'Binnen':''}, index=[0])
new_row6 = pd.DataFrame({'Voertuig':'BEDRIJFSAFVAL', 'Route':'', 'Chauffeur':'', 'Milieuwerker 1':'', 'Milieuwerker 2':'', 'Opmerkingen':'', 'OK/NOK':'','Gedaan':'', 'Binnen':''}, index=[0])
new_row7 = pd.DataFrame({'Voertuig':'CONTAINERTRUCKS', 'Route':'', 'Chauffeur':'', 'Milieuwerker 1':'', 'Milieuwerker 2':'', 'Opmerkingen':'', 'OK/NOK':'','Gedaan':'', 'Binnen':''}, index=[0])
new_row8 = pd.DataFrame({'Voertuig':'ONDERGRONDS', 'Route':'', 'Chauffeur':'', 'Milieuwerker 1':'', 'Milieuwerker 2':'', 'Opmerkingen':'', 'OK/NOK':'','Gedaan':'', 'Binnen':''}, index=[0])
dagplanning = pd.concat([new_row,df.loc[:]])
dagplanning = pd.concat([new_row2,df.loc[:]])
dagplanning = pd.concat([new_row3,df.loc[:]])
dagplanning = pd.concat([new_row4,df.loc[:]])
dagplanning = pd.concat([new_row5,df.loc[:]])
dagplanning = pd.concat([new_row6,df.loc[:]])
dagplanning = pd.concat([new_row7,df.loc[:]])
dagplanning = pd.concat([new_row8,df.loc[:]]).reset_index(drop=True)

# CONCATENEREN VAN HEADERS EN FRACTIES

dagplanning = pd.concat([new_row, huisvuil, new_row2, diftar, new_row3, papier, new_row4, glas, new_row5, grof, new_row6, bedrijf, new_row7, trucks, new_row8, ondergronds], axis=0)

# OPSLAAN VOOR OPMAAK


root.filename = filedialog.asksaveasfilename(initialdir = "/",title = "Select file",filetypes = (("xlsx files","*. xlsx"),("all files","*.*")))
exportlocatie2 = root.filename


dagplanning.to_excel(exportlocatie2, startrow = 4, index=False)
wb = opx.load_workbook(exportlocatie2)
sh1 = wb['Sheet1']

# OPMAAK GROOTTE

for row in sh1.iter_rows():
    for cell in row:
        cell.font = Font(size=14)

# HEADER VOOR DATUM

locale.setlocale(locale.LC_TIME, 'nl_NL.UTF-8')
today = datetime.datetime.now()
week = today.isocalendar().week
fdate = today.strftime("%A %d/%m/%Y")

sh1['A1'] = fdate
sh1['A1'].font = Font(bold = True, size=26)
sh1['A1'].fill =  PatternFill(start_color="FFDB58", end_color="FFDB58", fill_type="solid")
sh1['A6'].font = Font(bold = True, size = 18)
sh1['A6'].fill =  PatternFill(start_color="FCF55F", end_color="FCF55F", fill_type="solid")

# ALL BORDERS

for row in sh1.iter_rows():
    for cell in row:
        cell.border = Border(top = Side(border_style='thin', color='FF000000'),
                              right = Side(border_style='thin', color='FF000000'),
                              bottom = Side(border_style='thin', color='FF000000'),
                              left = Side(border_style='thin', color='FF000000'))

# KLEURENOPMAAK

for row in sh1.iter_rows():
    for cell in row:
        if cell.value == "START PMD: ":
            cell.font = Font(bold=True, size=14)
            varpmd = cell.row
        if cell.value == 'DIFTAR' and week%2 !=0:
            cell.value = 'DIFTAR//GFT'
            vardif = cell.row
            cell.font = Font(bold = True, size = 18)
            cell.fill = PatternFill(start_color="B4C424", end_color="B4C424", fill_type="solid")
        if cell.value == 'DIFTAR' and week%2 == 0:
            cell.value = 'DIFTAR//REST'
            vardif = cell.row
            cell.font = Font(bold = True, size=18)
            cell.fill = PatternFill(start_color= "C4B454", end_color= "C4B454", fill_type = "solid")
        if cell.value == 'PAPIER':
            varpap = cell.row
            cell.font = Font(bold = True, size = 18)
            cell.fill = PatternFill(start_color="FBCEB1", end_color="FBCEB1", fill_type="solid")
        if cell.value == 'GLAS':
            vargls = cell.row
            cell.font = Font(bold = True, size = 18)
            cell.fill = PatternFill(start_color="FFAA33",end_color="FFAA33", fill_type="solid")
        if cell.value == 'GROF VUIL':
            vargrfv = cell.row
            cell.font = Font(bold = True, size = 18)
            cell.fill = PatternFill(start_color="EEDC82",end_color="EEDC82", fill_type="solid")
        if cell.value == 'BEDRIJFSAFVAL':
            varba = cell.row
            cell.font = Font(bold = True, size = 18)
        if cell.value == 'CONTAINERTRUCKS':
            varct = cell.row
            cell.font = Font(bold = True, size=18)
        if cell.value == 'ONDERGRONDS':
            varoc = cell.row
            cell.font = Font(bold = True, size=18)
        else:
            pass

# FRACTIES ANDERE DIENSTEN GRIJS KLEUREN

lastrow = sh1.max_row
for rows in sh1.iter_rows(min_row=varba, max_row=lastrow, min_col=0, max_col=9):
    for cell in rows:
        cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type="solid")

# MERGING HEADERS AND CENTERING

sh1.merge_cells('A1:I4')
sh1.merge_cells('A6:I6')
sh1.merge_cells(start_row=vardif, end_row=vardif, start_column=1, end_column=9)
sh1.merge_cells(start_row=varpap, end_row=varpap, start_column=1, end_column=9)
sh1.merge_cells(start_row=vargls, end_row=vargls, start_column=1, end_column=9)
sh1.merge_cells(start_row=vargrfv, end_row=vargrfv, start_column=1, end_column=9)
sh1.merge_cells(start_row=varba, end_row=varba, start_column=1, end_column=9)
sh1.merge_cells(start_row=varct, end_row=varct, start_column=1, end_column=9)
sh1.merge_cells(start_row=varoc, end_row=varoc, start_column=1, end_column=9)

for row in range(1, sh1.max_row+1):
    for col in range(1, sh1.max_column+1):
        cell = sh1.cell(row, col)
        cell.alignment = Alignment(horizontal='center', vertical='center')

# SIZE OF COLUMNS

sh1.column_dimensions['A'].width = 10
sh1.column_dimensions['B'].width = 20
sh1.column_dimensions['C'].width = 30
sh1.column_dimensions['D'].width = 30
sh1.column_dimensions['E'].width = 30
sh1.column_dimensions['F'].width = 35
sh1.column_dimensions['G'].width = 15
sh1.column_dimensions['H'].width = 15
sh1.column_dimensions['I'].width = 15


wb.save(exportlocatie2)

exit_button = Button(root, text="Exit", command=root.destroy)
exit_button.pack(pady=20)
root.mainloop()










